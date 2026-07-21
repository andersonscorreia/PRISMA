import csv
import io
import json
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.utils import timezone
from django.db import models
from django.db.models import Q

from core.models import (
    Cliente, Brand, PrinterOID, Impressora, APIToken, ColetaImpressora, ComputadorAgente,
    StatusImpressora, OrigemContador, HistoricoMovimentacao, HistoricoContador
)
from core.forms import ClienteForm, ImpressoraForm, UserRegistrationForm, UserEditForm, PerfilOidMarcaForm, BrandForm, PrinterOIDForm, PrinterStockForm
from core.serializers import (
    validar_coleta_agente_payload, serialize_impressora, 
    serialize_historico_movimentacao, serialize_historico_contador
)
from core.services import alterar_status_impressora, tem_colunas_subcontadores

# =====================================================
# DECORADORES E CONTROLE DE ACESSO
# =====================================================
def group_required(*group_names):
    def check(user):
        if not user.is_authenticated:
            return False
        if user.is_superuser or user.is_staff or user.groups.filter(name__in=group_names).exists() or not user.groups.exists():
            return True
        raise PermissionDenied
    return user_passes_test(check, login_url='login')

# =====================================================
# AUTENTICAÇÃO
# =====================================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard_geral')
        
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(username=u, password=p)
        if user is not None:
            login(request, user)
            return redirect('dashboard_geral')
        else:
            messages.error(request, 'Usuário ou senha incorretos.')
            
    return render(request, 'core/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

# =====================================================
# WEBPAGES / DASHBOARD E MÉTRICAS
# =====================================================
@login_required(login_url='login')
def dashboard_geral(request):
    cliente_id = request.GET.get('cliente_id')
    clientes = Cliente.objects.all().order_by('nome')
    
    # Mostra todas as impressoras alocadas (com ou sem IP)
    impressoras = Impressora.objects.filter(status='ALOCADA').select_related('cliente', 'brand')
    
    if cliente_id:
        cliente_filtrado = get_object_or_404(Cliente, pk=cliente_id)
        impressoras = impressoras.filter(cliente=cliente_filtrado)
        
        total_locadas    = Impressora.objects.filter(cliente=cliente_filtrado, status='ALOCADA').count()
        total_disponiveis = Impressora.objects.filter(cliente=cliente_filtrado, status='ESTOQUE').count()
        total_assistencia = Impressora.objects.filter(cliente=cliente_filtrado, status_sistema='Inativo').count()
    else:
        cliente_filtrado  = None
        total_locadas     = Impressora.objects.filter(status='ALOCADA').count()
        total_disponiveis = Impressora.objects.filter(status='ESTOQUE').count()
        total_assistencia = Impressora.objects.filter(status_sistema='Inativo').count()

    alocacoes = []
    for imp in impressoras:
        alocacoes.append({
            'alocacao': {
                'impressora': imp,
                'cliente': imp.cliente,
                'data_entrada': imp.data_alocacao or date.today(),
                'observacoes': 'Ativo Alocado'
            }
        })

    # ---- Preview de contadores para a tabela na página ----
    data_inicio_str = request.GET.get('data_inicio', '').strip()
    data_fim_str    = request.GET.get('data_fim', request.GET.get('data_referencia', '')).strip()
    serial_query    = request.GET.get('serial', '').strip()
    tipo_filtro     = request.GET.get('tipo', '').strip()

    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else date.today()
    except (ValueError, TypeError):
        data_fim = date.today()

    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else date(data_fim.year, data_fim.month, 1)
    except (ValueError, TypeError):
        data_inicio = date(data_fim.year, data_fim.month, 1)

    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    # Queryset base de impressoras para preview (todas, filtradas por cliente/serial se selecionado)
    preview_qs = Impressora.objects.select_related('cliente', 'brand', 'oid_profile').order_by('cliente__nome', 'name')
    if cliente_id:
        preview_qs = preview_qs.filter(cliente__pk=cliente_id)
    if serial_query:
        preview_qs = preview_qs.filter(
            Q(serial_number__icontains=serial_query) |
            Q(modelo__icontains=serial_query) |
            Q(name__icontains=serial_query)
        )

    counter_rows = obter_linhas_contadores_discriminados(preview_qs, data_inicio=data_inicio, data_fim=data_fim)
    if tipo_filtro:
        counter_rows = [r for r in counter_rows if r['tipo'].lower() == tipo_filtro.lower()]

    # Calculo dos totais acumulados no período por tipo de papel
    soma_periodo_total = sum(r['consumo'] for r in counter_rows)
    soma_a4    = sum(r['consumo'] for r in counter_rows if r['tipo'] == 'A4')
    soma_a3    = sum(r['consumo'] for r in counter_rows if r['tipo'] == 'A3')
    soma_a5    = sum(r['consumo'] for r in counter_rows if r['tipo'] == 'A5')
    soma_pb    = sum(r['consumo'] for r in counter_rows if r['tipo'] == 'Total')
    soma_color = sum(r['consumo'] for r in counter_rows if r['tipo'] == 'Color')

    # Pré-calcular selected dos selects para evitar comparações == no template
    clientes_opcoes = [
        {'id': c.id, 'nome': c.nome, 'selected': (cliente_filtrado and c.id == cliente_filtrado.id)}
        for c in clientes
    ]
    tipos_opcoes = [
        {'valor': v, 'label': v, 'selected': (tipo_filtro == v)}
        for v in ['Total', 'A3', 'A4', 'A5']
    ]

    context = {
        'clientes': clientes,
        'clientes_opcoes': clientes_opcoes,
        'tipos_opcoes': tipos_opcoes,
        'cliente_filtrado': cliente_filtrado,
        'alocacoes': alocacoes,
        'total_locadas': total_locadas,
        'total_disponiveis': total_disponiveis,
        'total_assistencia': total_assistencia,
        'counter_rows': counter_rows,
        'data_inicio': data_inicio.strftime('%Y-%m-%d'),
        'data_inicio_br': data_inicio.strftime('%d/%m/%Y'),
        'data_fim': data_fim.strftime('%Y-%m-%d'),
        'data_fim_br': data_fim.strftime('%d/%m/%Y'),
        'data_referencia': data_fim.strftime('%Y-%m-%d'),
        'data_referencia_br': data_fim.strftime('%d/%m/%Y'),
        'serial_query': serial_query,
        'tipo_filtro': tipo_filtro,
        'soma_periodo_total': soma_periodo_total,
        'soma_periodo_total_fmt': f"{soma_periodo_total:,}".replace(',', '.'),
        'soma_a4_fmt': f"{soma_a4:,}".replace(',', '.'),
        'soma_a3_fmt': f"{soma_a3:,}".replace(',', '.'),
        'soma_a5_fmt': f"{soma_a5:,}".replace(',', '.'),
        'soma_pb_fmt': f"{soma_pb:,}".replace(',', '.'),
        'soma_color_fmt': f"{soma_color:,}".replace(',', '.'),
    }
    return render(request, 'core/dashboard_geral.html', context)


def garantir_coleta_diaria(imp, dt=None, novos_valores=None):
    """
    Garante que a impressora possua um registro em HistoricoContador para a data 'dt' (padrão: hoje):
    - Desde o início do dia a impressora já possui o contador diário inicializado (herdando dos últimos conhecidos ou inicial).
    - À medida que ocorrem coletas no dia, atualiza o registro do dia com a leitura mais recente.
    """
    if dt is None:
        dt = timezone.localdate()
    now = timezone.now()

    col = imp.ultima_coleta

    if novos_valores:
        cpb = novos_valores.get('pb')
        ccolor = novos_valores.get('color', 0)
        ca3 = novos_valores.get('a3')
        ca4 = novos_valores.get('a4')
        ca5 = novos_valores.get('a5')
    else:
        ca3 = col.contador_a3 if (col and col.contador_a3 is not None) else None
        ca4 = col.contador_a4 if (col and col.contador_a4 is not None) else None
        ca5 = col.contador_a5 if (col and col.contador_a5 is not None) else None
        cpb = (col.contador_geral or col.contador_total) if (col and (col.contador_geral is not None or col.contador_total is not None)) else (ca4 if (ca4 is not None) else (imp.ultimo_contador_pb or (imp.contador_inicial or 0)))
        ccolor = ca3 if (ca3 is not None) else (imp.ultimo_contador_color or 0)

    if (ca5 is None or ca5 == 0) and cpb is not None and ca4 is not None and ca3 is not None:
        if cpb > (ca4 + ca3):
            ca5 = cpb - (ca4 + ca3)

    try:
        if tem_colunas_subcontadores():
            hist = HistoricoContador.objects.filter(
                impressora=imp,
                data_coleta=dt,
                origem=OrigemContador.DIARIO
            ).order_by('-timestamp').first()
        else:
            hist = HistoricoContador.objects.only(
                'id', 'impressora', 'data_coleta', 'timestamp', 'contador_pb', 'contador_color', 'origem'
            ).filter(
                impressora=imp,
                data_coleta=dt,
                origem=OrigemContador.DIARIO
            ).order_by('-timestamp').first()

        if not hist:
            if tem_colunas_subcontadores():
                hist = HistoricoContador.objects.create(
                    impressora=imp,
                    data_coleta=dt,
                    origem=OrigemContador.DIARIO,
                    timestamp=now,
                    contador_pb=cpb or 0,
                    contador_color=ccolor or 0,
                    contador_a3=ca3,
                    contador_a4=ca4,
                    contador_a5=ca5,
                )
            else:
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "INSERT INTO historico_contador (impressora_id, data_coleta, timestamp, contador_pb, contador_color, origem) VALUES (%s, %s, %s, %s, %s, %s)",
                        [imp.pk, dt, now, cpb or 0, ccolor or 0, OrigemContador.DIARIO]
                    )
                hist = HistoricoContador.objects.only(
                    'id', 'impressora', 'data_coleta', 'timestamp', 'contador_pb', 'contador_color', 'origem'
                ).filter(
                    impressora=imp,
                    data_coleta=dt,
                    origem=OrigemContador.DIARIO
                ).order_by('-timestamp').first()
        elif novos_valores:
            if cpb is not None: hist.contador_pb = cpb
            if ccolor is not None: hist.contador_color = ccolor
            if tem_colunas_subcontadores():
                if ca3 is not None: hist.contador_a3 = ca3
                if ca4 is not None: hist.contador_a4 = ca4
                if ca5 is not None: hist.contador_a5 = ca5
            hist.timestamp = now
            if tem_colunas_subcontadores():
                hist.save()
            else:
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "UPDATE historico_contador SET contador_pb = %s, contador_color = %s, timestamp = %s WHERE id = %s",
                        [hist.contador_pb, hist.contador_color, hist.timestamp, hist.pk]
                    )
        return hist
    except Exception as e:
        import logging
        logging.error(f"Erro em garantir_coleta_diaria: {e}")
        return None


def obter_linhas_contadores_discriminados(impressoras_qs, data_inicio=None, data_fim=None, data_ref=None):
    """
    Desmembra os contadores de cada impressora em linhas individuais no período selecionado (data_inicio até data_fim):
    - Se a impressora possui subcontadores A3, A4 ou A5 (ex: modelos Epson), cria 1 linha para cada tipo.
    - Se for impressora sem subcontadores por tamanho, cria 1 linha para Total e 1 para Color (se colorida).
    - Para cada tipo, calcula a leitura inicial, leitura final e a diferença (consumo) no período.
    """
    if data_fim is None:
        data_fim = data_ref or date.today()
    if data_inicio is None:
        data_inicio = date(data_fim.year, data_fim.month, 1)

    if isinstance(data_inicio, str):
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        except Exception:
            data_inicio = date(data_fim.year, data_fim.month, 1)

    if isinstance(data_fim, str):
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        except Exception:
            data_fim = date.today()

    counter_rows = []

    def buscar_historico_seguro(imp, dt_limite):
        # Garante que desde o início do dia existe registro para o dia solicitado
        garantir_coleta_diaria(imp, dt=dt_limite)
        try:
            hc = (
                HistoricoContador.objects
                .filter(impressora=imp, data_coleta__lte=dt_limite)
                .order_by('-data_coleta', '-timestamp')
                .first()
            )
            if hc:
                _ = hc.contador_a4
            return hc
        except Exception:
            try:
                return (
                    HistoricoContador.objects
                    .only('id', 'impressora', 'data_coleta', 'timestamp', 'contador_pb', 'contador_color', 'origem')
                    .filter(impressora=imp, data_coleta__lte=dt_limite)
                    .order_by('-data_coleta', '-timestamp')
                    .first()
                )
            except Exception:
                return None

    def extrair_valores_dict(hc, imp):
        hc_dict = hc.__dict__ if hc else {}
        a3 = hc_dict.get('contador_a3')
        a4 = hc_dict.get('contador_a4')
        a5 = hc_dict.get('contador_a5')
        pb = hc_dict.get('contador_pb') if ('contador_pb' in hc_dict and hc_dict['contador_pb'] is not None) else (imp.ultimo_contador_pb or 0)
        color = hc_dict.get('contador_color') if ('contador_color' in hc_dict and hc_dict['contador_color'] is not None) else (imp.ultimo_contador_color or 0)
        data_col = hc_dict.get('data_coleta')
        return {
            'a3': a3,
            'a4': a4,
            'a5': a5,
            'pb': pb,
            'color': color,
            'data_coleta': data_col
        }

    for imp in impressoras_qs:
        col = imp.ultima_coleta

        hc_fim = buscar_historico_seguro(imp, data_fim)
        val_fim_dict = extrair_valores_dict(hc_fim, imp)

        hc_inicio = buscar_historico_seguro(imp, data_inicio)
        val_inicio_dict = extrair_valores_dict(hc_inicio, imp) if hc_inicio else val_fim_dict

        a3_fim = val_fim_dict['a3'] if (val_fim_dict['a3'] is not None) else (col.contador_a3 if (col and col.contador_a3 is not None) else None)
        a4_fim = val_fim_dict['a4'] if (val_fim_dict['a4'] is not None) else (col.contador_a4 if (col and col.contador_a4 is not None) else None)
        a5_fim = val_fim_dict['a5'] if (val_fim_dict['a5'] is not None) else (col.contador_a5 if (col and col.contador_a5 is not None) else None)

        a3_ini = val_inicio_dict['a3'] if (val_inicio_dict['a3'] is not None) else a3_fim
        a4_ini = val_inicio_dict['a4'] if (val_inicio_dict['a4'] is not None) else a4_fim
        a5_ini = val_inicio_dict['a5'] if (val_inicio_dict['a5'] is not None) else a5_fim

        pb_fim = val_fim_dict['pb']
        color_fim = val_fim_dict['color']

        pb_ini = val_inicio_dict['pb'] if (val_inicio_dict['pb'] is not None) else pb_fim
        color_ini = val_inicio_dict['color'] if (val_inicio_dict['color'] is not None) else color_fim

        data_coleta_val = val_fim_dict['data_coleta']
        if data_coleta_val:
            coleta_str = data_coleta_val.strftime('%d/%m/%Y')
        elif col and col.data_coleta:
            coleta_str = col.data_coleta.strftime('%d/%m/%Y')
        else:
            coleta_str = 'Sem coleta'

        items = []
        tem_subcontadores = (a3_fim is not None) or (a4_fim is not None) or (a5_fim is not None)

        if tem_subcontadores:
            if a3_fim is not None:
                items.append({'tipo': 'A3', 'val_inicio': int(a3_ini or 0), 'val_fim': int(a3_fim or 0)})
            if a4_fim is not None:
                items.append({'tipo': 'A4', 'val_inicio': int(a4_ini or 0), 'val_fim': int(a4_fim or 0)})
            if a5_fim is not None:
                items.append({'tipo': 'A5', 'val_inicio': int(a5_ini or 0), 'val_fim': int(a5_fim or 0)})
        else:
            items.append({'tipo': 'Total', 'val_inicio': int(pb_ini or 0), 'val_fim': int(pb_fim or 0)})
            if color_fim > 0 or color_ini > 0:
                items.append({'tipo': 'Color', 'val_inicio': int(color_ini or 0), 'val_fim': int(color_fim or 0)})

        equip_nome = f"{imp.brand.name} {imp.modelo or imp.name or ''}".strip() if imp.brand else (imp.modelo or imp.name or '—')

        for item in items:
            v_ini = item['val_inicio']
            v_fim = item['val_fim']
            consumo = max(0, v_fim - v_ini)

            counter_rows.append({
                'serial': imp.serial_number or '—',
                'modelo': imp.modelo or imp.name or '—',
                'marca': imp.brand.name if imp.brand else '—',
                'equipamento': equip_nome,
                'cliente': imp.cliente.nome if imp.cliente else '—',
                'ip': imp.ip_address or '—',
                'tipo': item['tipo'],
                'valor_inicio': v_ini,
                'valor_inicio_formatado': f"{v_ini:,}".replace(',', '.'),
                'valor_fim': v_fim,
                'valor_fim_formatado': f"{v_fim:,}".replace(',', '.'),
                'valor': v_fim,  # compatibilidade
                'valor_formatado': f"{v_fim:,}".replace(',', '.'),
                'consumo': consumo,
                'consumo_formatado': f"{consumo:,}".replace(',', '.'),
                'coleta': coleta_str,
                'pk': imp.pk,
            })
    return counter_rows


@login_required(login_url='login')
def historico_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    
    historicos = []
    for imp in Impressora.objects.filter(cliente=cliente):
        historicos.append({
            'impressora': imp,
            'data_entrada': date.today(),
            'data_saida': None,
            'observacoes': 'Instalação ativa'
        })
        
    context = {
        'cliente': cliente,
        'historicos': historicos,
    }
    return render(request, 'core/historico_cliente.html', context)

@login_required(login_url='login')
def ciclo_vida_ativo(request, pk):
    impressora = get_object_or_404(Impressora, pk=pk)
    garantir_coleta_diaria(impressora)

    if not HistoricoContador.objects.filter(impressora=impressora).exclude(data_coleta=timezone.localdate()).exists():
        dt_base = impressora.data_alocacao.date() if impressora.data_alocacao else (impressora.updated_at.date() if impressora.updated_at else timezone.localdate())
        garantir_coleta_diaria(impressora, dt=dt_base)
    
    primeira_mov = HistoricoMovimentacao.objects.filter(impressora=impressora).order_by('data_movimentacao').first()
    try:
        primeira_col = HistoricoContador.objects.only('data_coleta').filter(impressora=impressora).order_by('data_coleta').first()
    except Exception:
        primeira_col = None
    
    datas_possiveis = []
    if primeira_mov and primeira_mov.data_movimentacao:
        datas_possiveis.append(primeira_mov.data_movimentacao.date())
    if primeira_col and primeira_col.data_coleta:
        datas_possiveis.append(primeira_col.data_coleta)
    if impressora.data_alocacao:
        datas_possiveis.append(impressora.data_alocacao.date())
    if impressora.updated_at:
        datas_possiveis.append(impressora.updated_at.date())
        
    data_limite_cadastro = min(datas_possiveis) if datas_possiveis else timezone.localdate()
    data_cadastro_str = data_limite_cadastro.strftime('%Y-%m-%d')
    data_cadastro_br = data_limite_cadastro.strftime('%d/%m/%Y')
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    def parse_dt(val):
        if not val:
            return None
        val_str = str(val).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                pass
        return None

    dt_ini = parse_dt(data_inicio)
    dt_fim = parse_dt(data_fim)

    try:
        historicos_contador = HistoricoContador.objects.filter(impressora=impressora).order_by('-data_coleta', '-timestamp')
        if dt_ini:
            historicos_contador = historicos_contador.filter(data_coleta__gte=dt_ini)
        if dt_fim:
            historicos_contador = historicos_contador.filter(data_coleta__lte=dt_fim)
        todos_contadores = list(historicos_contador)
    except Exception:
        try:
            historicos_contador = HistoricoContador.objects.only(
                'id', 'impressora', 'data_coleta', 'timestamp', 'contador_pb', 'contador_color', 'origem'
            ).filter(impressora=impressora).order_by('-data_coleta', '-timestamp')
            if dt_ini:
                historicos_contador = historicos_contador.filter(data_coleta__gte=dt_ini)
            if dt_fim:
                historicos_contador = historicos_contador.filter(data_coleta__lte=dt_fim)
            todos_contadores = list(historicos_contador)
        except Exception:
            todos_contadores = []

    historicos_movimentacao = HistoricoMovimentacao.objects.filter(
        impressora=impressora
    ).select_related('cliente').order_by('-data_movimentacao')

    if dt_ini:
        historicos_movimentacao = historicos_movimentacao.filter(data_movimentacao__date__gte=dt_ini)
    if dt_fim:
        historicos_movimentacao = historicos_movimentacao.filter(data_movimentacao__date__lte=dt_fim)
    
    # Manter todas as coletas diárias (mesmo se o valor for o mesmo do dia anterior)
    movs_list = list(historicos_movimentacao)
    import json as _json
    mov_intervals = []
    now_iso = timezone.now().isoformat()
    for idx, mov in enumerate(movs_list):
        inicio_iso = mov.data_movimentacao.isoformat()
        if idx == 0:
            fim_iso = now_iso
        else:
            fim_iso = movs_list[idx - 1].data_movimentacao.isoformat()
        mov_intervals.append({'inicio': inicio_iso, 'fim': fim_iso})

    context = {
        'impressora': impressora,
        'historicos_movimentacao': movs_list,
        'historicos_contador': todos_contadores,
        'mov_intervals_json': _json.dumps(mov_intervals),
        'data_inicio': data_inicio or '',
        'data_fim': data_fim or '',
        'data_cadastro_str': data_cadastro_str,
        'data_cadastro_br': data_cadastro_br,
    }
    return render(request, 'core/ciclo_vida.html', context)


@login_required(login_url='login')
def estoque_disponivel(request):
    impressoras = Impressora.objects.filter(status=StatusImpressora.ESTOQUE).select_related('brand')
    return render(request, 'core/estoque.html', {'impressoras': impressoras})

@login_required(login_url='login')
def movimentacao_impressoras(request):
    impressoras = Impressora.objects.all()
    clientes = Cliente.objects.all()
    
    if request.method == 'POST':
        messages.success(request, "Simulação de movimentação concluída com sucesso.")
        return redirect('movimentacao_impressoras')
        
    return render(request, 'core/movimentacao.html', {
        'impressoras': impressoras,
        'clientes': clientes
    })


@login_required(login_url='login')
def exportar_historico_counters(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="historico_contadores.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Nº de Série', 'Modelo', 'Marca', 'Cliente', 'IP', 'Contador PB', 'Contador Color', 'Data da Coleta'])
    
    historicos = HistoricoContador.objects.select_related('impressora', 'impressora__cliente', 'impressora__brand').all().order_by('-data_coleta')
    for h in historicos:
        imp = h.impressora
        writer.writerow([
            imp.serial_number,
            imp.modelo,
            imp.brand.name if imp.brand else '',
            imp.cliente.nome if imp.cliente else '',
            imp.ip_address,
            h.contador_pb,
            h.contador_color,
            h.data_coleta.strftime('%d/%m/%Y')
        ])
        
    return response


@login_required(login_url='login')
def exportar_contadores_xlsx(request):
    """
    Exporta, para cada impressora alocada, os contadores do período especificado (data_inicio até data_fim).
    Colunas: Equipamento, Tipo de Contador, Cliente, Endereço IP, Nº de Série, Leitura Inicial, Leitura Final, Produção no Período, Data da Coleta.
    """
    data_inicio_str = request.GET.get('data_inicio', '').strip()
    data_fim_str    = request.GET.get('data_fim', request.GET.get('data_referencia', '')).strip()
    cliente_id      = request.GET.get('cliente_id', '').strip()
    serial_query    = request.GET.get('serial', '').strip()
    tipo_filtro     = request.GET.get('tipo', '').strip()

    try:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else date.today()
    except (ValueError, TypeError):
        data_fim = date.today()

    try:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else date(data_fim.year, data_fim.month, 1)
    except (ValueError, TypeError):
        data_inicio = date(data_fim.year, data_fim.month, 1)

    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    impressoras_qs = Impressora.objects.select_related('cliente', 'brand', 'oid_profile').order_by('cliente__nome', 'name')
    if cliente_id:
        impressoras_qs = impressoras_qs.filter(cliente__pk=cliente_id)
    if serial_query:
        impressoras_qs = impressoras_qs.filter(
            Q(serial_number__icontains=serial_query) |
            Q(modelo__icontains=serial_query) |
            Q(name__icontains=serial_query)
        )

    linhas = obter_linhas_contadores_discriminados(impressoras_qs, data_inicio=data_inicio, data_fim=data_fim)
    if tipo_filtro:
        linhas = [r for r in linhas if r['tipo'].lower() == tipo_filtro.lower()]

    soma_periodo_total = sum(r['consumo'] for r in linhas)
    soma_a4    = sum(r['consumo'] for r in linhas if r['tipo'] == 'A4')
    soma_a3    = sum(r['consumo'] for r in linhas if r['tipo'] == 'A3')
    soma_a5    = sum(r['consumo'] for r in linhas if r['tipo'] == 'A5')
    soma_pb    = sum(r['consumo'] for r in linhas if r['tipo'] == 'Total')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contadores por Período'

    header_fill   = PatternFill('solid', fgColor='C00000')          # vermelho PRISMA
    header_font   = Font(bold=True, color='FFFFFF', size=11)
    summary_bg    = PatternFill('solid', fgColor='F9F9F9')
    total_fill    = PatternFill('solid', fgColor='FFF2F2')
    thin          = Side(style='thin', color='D0D0D0')
    thick_bottom  = Side(style='double', color='C00000')
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_border  = Border(left=thin, right=thin, top=thin, bottom=thick_bottom)
    center        = Alignment(horizontal='center', vertical='center')
    left          = Alignment(horizontal='left',   vertical='center')
    right         = Alignment(horizontal='right',  vertical='center')

    # 1. Título
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'Relatório de Contadores por Período — {data_inicio.strftime("%d/%m/%Y")} até {data_fim.strftime("%d/%m/%Y")}'
    title_cell.font  = Font(bold=True, size=13, color='C00000')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # 2. Sumário de Totais do Início
    ws.merge_cells('A3:I3')
    sum_title = ws['A3']
    sum_title.value = "SUMÁRIO DE PRODUÇÃO ACUMULADA NO PERÍODO"
    sum_title.font = Font(bold=True, size=10, color='404040')
    sum_title.alignment = left

    sum_items = [
        ('PRODUÇÃO TOTAL', soma_periodo_total),
        ('PRODUÇÃO A4', soma_a4),
        ('PRODUÇÃO A3', soma_a3),
        ('PRODUÇÃO A5', soma_a5),
        ('PRODUÇÃO TOTAL (PB)', soma_pb),
    ]

    col_positions = [1, 3, 5, 7, 9]
    for idx, (label, val) in enumerate(sum_items):
        c_pos = col_positions[idx]
        lbl_cell = ws.cell(row=4, column=c_pos, value=f"{label}:")
        lbl_cell.font = Font(bold=True, size=9, color='666666')
        lbl_cell.alignment = right
        
        val_cell = ws.cell(row=4, column=c_pos+1 if c_pos < 9 else c_pos, value=val if c_pos < 9 else f"{label}: {val:,}".replace(',', '.'))
        if c_pos < 9:
            val_cell.font = Font(bold=True, size=10, color='C00000')
            val_cell.number_format = '#,##0'
            val_cell.alignment = left

    # 3. Cabeçalho da Tabela
    headers = ['Equipamento', 'Tipo de Contador', 'Cliente', 'Endereço IP', 'Nº de Série', 'Leitura Inicial', 'Leitura Final', 'Produção no Período', 'Data da Coleta']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = cell_border
    ws.row_dimensions[6].height = 22

    # 4. Dados
    for row_idx, linha in enumerate(linhas, start=7):
        valores = [
            linha['equipamento'],
            linha['tipo'],
            linha['cliente'],
            linha['ip'],
            linha['serial'],
            linha['valor_inicio'],
            linha['valor_fim'],
            linha['consumo'],
            linha['coleta'],
        ]
        for col_idx, v in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = cell_border
            if col_idx in (6, 7, 8):  # Leitura Inicial, Leitura Final, Produção no Período
                cell.alignment = right
                cell.number_format = '#,##0'
                if col_idx == 8:
                    cell.font = Font(bold=True)
            elif col_idx in (2, 4, 5, 9): # Tipo, IP, Serial, Coleta
                cell.alignment = center
            else:
                cell.alignment = left

    # 5. Linha de Total Geral ao final da Tabela
    total_row_idx = len(linhas) + 7
    ws.cell(row=total_row_idx, column=1, value="TOTAL GERAL").font = Font(bold=True, size=11, color='C00000')
    ws.cell(row=total_row_idx, column=1).alignment = left
    ws.cell(row=total_row_idx, column=1).border = total_border
    ws.cell(row=total_row_idx, column=1).fill = total_fill

    for c_idx in range(2, 9):
        cell = ws.cell(row=total_row_idx, column=c_idx)
        cell.border = total_border
        cell.fill = total_fill
        if c_idx == 8:
            cell.value = soma_periodo_total
            cell.font = Font(bold=True, size=11, color='C00000')
            cell.alignment = right
            cell.number_format = '#,##0'

    col_widths = [24, 18, 26, 16, 20, 16, 16, 20, 16]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- rodapé --
    footer_row = total_row_idx + 2
    ws.merge_cells(f'A{footer_row}:I{footer_row}')
    footer_cell = ws[f'A{footer_row}']
    footer_cell.value = f'Gerado pelo sistema PRISMA em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    footer_cell.font  = Font(italic=True, size=9, color='888888')
    footer_cell.alignment = center

    # -- stream resposta --
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f'contadores_{data_inicio.strftime("%Y%m%d")}_a_{data_fim.strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


@login_required(login_url='login')
def exportar_historico_counters(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="historico_contadores.csv"'
    response.write(u'\ufeff'.encode('utf8'))
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Data/Hora', 'IP da Impressora', 'Fabricante', 'Modelo', 'Contador Geral'])
    
    try:
        coletas = ColetaImpressora.objects.all().order_by('ip')
        for col in coletas:
            time_str = col.data_coleta.strftime('%d/%m/%Y %H:%M:%S') if col.data_coleta else 'N/A'
            db_imp = Impressora.objects.filter(ip_address=col.ip).first()
            brand = db_imp.brand.name if db_imp else 'Generic'
            
            writer.writerow([
                time_str,
                col.ip,
                brand,
                col.modelo or 'N/A',
                col.contador_geral or col.contador_total or 0
            ])
    except Exception as e:
        writer.writerow([f"Erro ao consultar MySQL: {str(e)}"])
        
    return response

# =====================================================
# PAINEL DE CADASTROS ADMINISTRATIVOS (Admin Only)
# =====================================================
@login_required(login_url='login')
@group_required('Admin')
def gerenciamento_painel(request):
    clientes = Cliente.objects.all().order_by('nome')
    impressoras = Impressora.objects.all().order_by('name')
    usuarios = User.objects.all().order_by('username')
    perfis_oid = PrinterOID.objects.all().order_by('brand__name')
    marcas = Brand.objects.all().order_by('name')
    
    return render(request, 'core/gerenciamento.html', {
        'clientes': clientes,
        'impressoras': impressoras,
        'usuarios': usuarios,
        'perfis_oid': perfis_oid,
        'marcas': marcas
    })

# --- CADASTROS ---
@login_required(login_url='login')
@group_required('Admin')
def cadastrar_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            APIToken.objects.create(
                cliente=cliente,
                nome_identificador=f"Token Principal - {cliente.nome}",
                tipo_acesso="full_access"
            )
            messages.success(request, "Cliente cadastrado com sucesso! APIToken gerado.")
            return redirect('gerenciamento')
    else:
        form = ClienteForm()
    return render(request, 'core/cadastro_cliente.html', {'form': form})

@login_required(login_url='login')
@group_required('Admin')
def cadastrar_impressora(request):
    if request.method == 'POST':
        form = ImpressoraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Impressora cadastrada com sucesso!")
            return redirect('gerenciamento')
    else:
        form = ImpressoraForm()
    return render(request, 'core/cadastro_impressora.html', {'form': form})

@login_required(login_url='login')
@group_required('Admin')
def cadastrar_usuario(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuário cadastrado com sucesso!")
            return redirect('gerenciamento')
    else:
        form = UserRegistrationForm()
    return render(request, 'core/cadastro_usuario.html', {'form': form})

@login_required(login_url='login')
@group_required('Admin')
def cadastrar_perfil_oid(request):
    if request.method == 'POST':
        form = PerfilOidMarcaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil de OID por Marca cadastrado com sucesso!")
            return redirect('gerenciamento')
    else:
        form = PerfilOidMarcaForm()
    return render(request, 'core/cadastro_perfil_oid.html', {'form': form})

@login_required(login_url='login')
@group_required('Admin')
def cadastrar_marca(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            brand = form.save()
            messages.success(request, f"Marca '{brand.name}' cadastrada com sucesso!")
            return redirect('gerenciamento')
    else:
        form = BrandForm()
    return render(request, 'core/cadastro_marca.html', {'form': form})

# --- EDIÇÕES ---
@login_required(login_url='login')
@group_required('Admin')
def editar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente atualizado com sucesso!")
            return redirect('gerenciamento')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'core/cadastro_cliente.html', {
        'form': form,
        'title': "Editar Cliente",
        'subtitle': f"Atualize os dados de {cliente.nome}."
    })

@login_required(login_url='login')
@group_required('Admin')
def editar_impressora(request, pk):
    impressora = get_object_or_404(Impressora, pk=pk)
    if request.method == 'POST':
        form = ImpressoraForm(request.POST, instance=impressora)
        if form.is_valid():
            form.save()
            messages.success(request, "Impressora atualizada com sucesso!")
            return redirect('gerenciamento')
    else:
        form = ImpressoraForm(instance=impressora)
    return render(request, 'core/cadastro_impressora.html', {
        'form': form,
        'title': "Editar Impressora",
        'subtitle': f"Atualize os dados de {impressora.name or impressora.serial_number}."
    })

@login_required(login_url='login')
@group_required('Admin')
def editar_usuario(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuário atualizado com sucesso!")
            return redirect('gerenciamento')
    else:
        form = UserEditForm(instance=user)
    return render(request, 'core/cadastro_usuario.html', {
        'form': form,
        'title': "Editar Usuário / Operador",
        'subtitle': f"Atualize os dados de {user.username}."
    })

@login_required(login_url='login')
@group_required('Admin')
def editar_perfil_oid(request, pk):
    perfil = get_object_or_404(PrinterOID, pk=pk)
    if request.method == 'POST':
        form = PerfilOidMarcaForm(request.POST, instance=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil de OID atualizado com sucesso!")
            return redirect('gerenciamento')
    else:
        form = PerfilOidMarcaForm(instance=perfil)
    return render(request, 'core/cadastro_perfil_oid.html', {
        'form': form,
        'title': "Editar Perfil de OID",
        'subtitle': f"Atualize os mapeamentos de OIDs para a marca {perfil.brand.name}."
    })

@login_required(login_url='login')
@group_required('Admin')
def editar_marca(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, f"Marca '{brand.name}' atualizada com sucesso!")
            return redirect('gerenciamento')
    else:
        form = BrandForm(instance=brand)
    return render(request, 'core/cadastro_marca.html', {
        'form': form,
        'title': "Editar Marca",
        'subtitle': f"Atualize o nome da marca {brand.name}."
    })

# --- EXCLUSÕES ---
@login_required(login_url='login')
@group_required('Admin')
def excluir_cliente(request, pk):
    cliente = Cliente.objects.filter(pk=pk).first()
    if cliente:
        nome = cliente.nome
        cliente.delete()
        messages.success(request, f"Cliente '{nome}' excluído com sucesso!")
    else:
        messages.info(request, "O cliente solicitado não foi encontrado ou já foi excluído.")
    return redirect('gerenciamento')

@login_required(login_url='login')
@group_required('Admin')
def excluir_impressora(request, pk):
    impressora = Impressora.objects.filter(pk=pk).first() or Impressora.objects.filter(serial_number__iexact=str(pk).strip()).first()
    if impressora:
        nome_imp = impressora.name or impressora.modelo or impressora.serial_number
        impressora.delete()
        messages.success(request, f"Impressora '{nome_imp}' ({pk}) excluída com sucesso!")
    else:
        messages.info(request, f"A impressora '{pk}' não foi encontrada ou já foi excluída.")
    return redirect('gerenciamento')

@login_required(login_url='login')
@group_required('Admin')
def excluir_usuario(request, pk):
    user = User.objects.filter(pk=pk).first()
    if not user:
        messages.info(request, "O usuário solicitado não foi encontrado ou já foi excluído.")
    elif user == request.user:
        messages.error(request, "Você não pode excluir a si mesmo.")
    else:
        username = user.username
        user.delete()
        messages.success(request, f"Usuário '{username}' excluído com sucesso!")
    return redirect('gerenciamento')

@login_required(login_url='login')
@group_required('Admin')
def excluir_perfil_oid(request, pk):
    perfil = PrinterOID.objects.filter(pk=pk).first()
    if perfil:
        brand_name = perfil.brand.name if perfil.brand else "Desconhecida"
        perfil.delete()
        messages.success(request, f"Perfil de OID da marca '{brand_name}' excluído com sucesso!")
    else:
        messages.info(request, "O perfil de OID não foi encontrado ou já foi excluído.")
    return redirect('gerenciamento')

@login_required(login_url='login')
@group_required('Admin')
def excluir_marca(request, pk):
    brand = Brand.objects.filter(pk=pk).first()
    if not brand:
        messages.info(request, "A marca solicitada não foi encontrada ou já foi excluída.")
        return redirect('gerenciamento')
        
    brand_name = brand.name
    try:
        brand.delete()
        messages.success(request, f"Marca '{brand_name}' excluída com sucesso!")
    except Exception:
        messages.error(request, f"Não é possível excluir a marca '{brand_name}' porque ela possui impressoras ou perfis de OID associados.")
    return redirect('gerenciamento')

# =====================================================
# CONTROLE E GESTÃO DE TOKENS FRONT-END
# =====================================================
# ENDPOINT DE SINCRONIZAÇÃO DA COLETA (API SEGURA)
# =====================================================
from django.views.decorators.http import require_GET

@csrf_exempt
@require_GET
def check_printer_config(request):
    """
    Endpoint GET para handshake do agente SNMP.
    Busca a impressora pelo número de série no MySQL e retorna seus OIDs associados.
    """
    serial_number = request.GET.get('serial_number')
    if not serial_number:
        return JsonResponse({"error": "Parâmetro 'serial_number' é obrigatório."}, status=400)
        
    try:
        printer = Impressora.objects.select_related('brand').get(serial_number=serial_number)
    except Impressora.DoesNotExist:
        return JsonResponse({"error": "Impressora não cadastrada no servidor."}, status=404)
        
    try:
        oid_config = printer.oid_profile or PrinterOID.objects.filter(brand=printer.brand).first() or PrinterOID.objects.filter(brands=printer.brand).first()
        if not oid_config:
            raise PrinterOID.DoesNotExist()
        oids_data = {
            "printer_oid": oid_config.printer_oid,
            "oid_serial_number": oid_config.oid_serial_number,
            "oid_tempo_ligada": oid_config.oid_tempo_ligada,
            "oid_mensagem_painel": oid_config.oid_mensagem_painel,
            "oid_counter_total": oid_config.oid_counter_total,
            "oid_counter_mono": oid_config.oid_counter_mono,
            "oid_counter_color": oid_config.oid_counter_color,
            "oid_toner_level": oid_config.oid_toner_level,
            "oid_toner_full": oid_config.oid_toner_full,
            "oid_tinta_preta": oid_config.oid_tinta_preta,
            "oid_tinta_ciano": oid_config.oid_tinta_ciano,
            "oid_tinta_magenta": oid_config.oid_tinta_magenta,
            "oid_tinta_amarela": oid_config.oid_tinta_amarela,
            "oid_caixa_manutencao": oid_config.oid_caixa_manutencao,
        }
    except PrinterOID.DoesNotExist:
        return JsonResponse(
            {"error": f"Nenhuma configuração de OID cadastrada para a marca '{printer.brand.name}'."}, 
            status=422
        )
        
    return JsonResponse({
        "status": "success",
        "printer": {
            "serial_number": printer.serial_number,
            "model": printer.modelo or printer.name or "Genérica",
            "brand": printer.brand.name,
            "ip_address": printer.ip_address,
        },
        "oids": oids_data
    }, status=200)


@csrf_exempt
def config_coleta_api(request):
    """
    API protegida por Bearer Token. Valida o token contra a tabela APIToken
    e exige escopo de oids_only ou full_access.
    """
    if request.method != 'GET':
        return JsonResponse({"error": "Método não permitido."}, status=405)
        
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if not auth_header.startswith('Bearer '):
        return JsonResponse({"error": "Autenticação Bearer obrigatória."}, status=401)
        
    token = auth_header.split('Bearer ', 1)[1].strip()
    
    try:
        api_token = APIToken.objects.select_related('cliente').get(token_chave=token)
    except APIToken.DoesNotExist:
        return JsonResponse({"error": "Token inválido."}, status=403)
        
    if not api_token.ativo:
        return JsonResponse({"error": "Token inativo."}, status=403)
        
    if api_token.tipo_acesso not in ['oids_only', 'full_access']:
        return JsonResponse(
            {"error": "Este token não possui permissão para acessar configurações de OIDs"}, 
            status=403
        )
        
    cliente = api_token.cliente
    impressora = Impressora.objects.filter(cliente=cliente, status_sistema='Ativo').first()
    if not impressora:
        return JsonResponse({"error": "Nenhuma impressora ativa."}, status=404)
        
    perfil = impressora.oid_profile or PrinterOID.objects.filter(brand=impressora.brand).first() or PrinterOID.objects.filter(brands=impressora.brand).first()
    oids_map = {}
    if perfil:
        oids_map = {
            "printer_oid": perfil.printer_oid,
            "contador": perfil.oid_counter_total,
            "tempo_ligada": "1.3.6.1.2.1.1.3.0",
            "serial": perfil.oid_serial_number,
            "mensagem_painel": "1.3.6.1.2.1.43.16.5.1.2.1.1",
            "toner_atual": perfil.oid_toner_level,
            "toner_full": "100",
        }
        
    return JsonResponse({
        "ip_address": impressora.ip_address,
        "nome_comercial": impressora.name or impressora.nome_comercial or "Sem Nome",
        "name": impressora.name or impressora.nome_comercial or "Sem Nome",
        "modelo": impressora.modelo or "Genérica",
        "model": impressora.modelo or "Genérica",
        "oids": oids_map
    })

# =====================================================
# ENDPOINT DE TELEMETRIA (GRAVAÇÃO INFLUXDB)
# =====================================================
@csrf_exempt
def api_metrics_insert(request):
    """
    Recebe telemetria do agente externo via POST.
    Autenticação por Bearer Token na tabela APIToken, exigindo escopo 'metrics_write' ou 'full_access'.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método não permitido. Utilize POST."}, status=405)
        
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if not auth_header.startswith('Bearer '):
        return JsonResponse({"error": "Autenticação Bearer obrigatória."}, status=401)
        
    token = auth_header.split('Bearer ', 1)[1].strip()
    try:
        api_token = APIToken.objects.get(token_chave=token)
    except APIToken.DoesNotExist:
        return JsonResponse({"error": "Token inválido."}, status=403)
        
    if not api_token.ativo:
        return JsonResponse({"error": "Token inativo."}, status=403)
        
    if api_token.tipo_acesso not in ['metrics_write', 'full_access']:
        return JsonResponse(
            {"error": "Este token não possui permissão para gravar métricas de telemetria."}, 
            status=403
        )
        
    try:
        data = json.loads(request.body)
        
        ip = data.get("ip_address")
        serial = data.get("serial_number")
        if not ip or not serial:
            return JsonResponse({"error": "ip_address e serial_number obrigatórios"}, status=400)
            
        serial_val = serial if serial and serial not in ("N/A", "---") else f"NO-SERIAL-{ip}"
        
        tinta_preta = None
        tinta_ciano = None
        tinta_magenta = None
        tinta_amarela = None
        caixa_manutencao = None
        
        toner_list = data.get("last_toner_data", []) or []
        for item in toner_list:
            color = item.get("color")
            level = item.get("level")
            try:
                level_val = float(level) if level not in ("N/A", "---") else None
            except (ValueError, TypeError):
                level_val = None
            if color == "Black":
                tinta_preta = level_val
            elif color == "Cyan":
                tinta_ciano = level_val
            elif color == "Magenta":
                tinta_magenta = level_val
            elif color == "Yellow":
                tinta_amarela = level_val
            elif color == "Manutenção":
                caixa_manutencao = level_val
                
        def parse_int(v):
            try:
                return int(v) if v is not None else None
            except (ValueError, TypeError):
                return None
                
        ColetaImpressora.objects.update_or_create(
            serial=serial_val,
            defaults={
                "ip": ip,
                "contador_geral": parse_int(data.get("last_counter")),
                "uptime": data.get("tempo_ligada"),
                "mensagem_painel": data.get("mensagem_erro"),
                "porcentagem_toner": tinta_preta,
                "status": data.get("status", "Online"),
                "modelo": data.get("model"),
                "tinta_preta": tinta_preta,
                "tinta_ciano": tinta_ciano,
                "tinta_magenta": tinta_magenta,
                "tinta_amarela": tinta_amarela,
                "caixa_manutencao": caixa_manutencao
            }
        )

        # Atualizar também o modelo Impressora e HistoricoContador (origem='DIARIO')
        if serial_val and not serial_val.startswith("NO-SERIAL"):
            cpb = parse_int(data.get("last_counter")) or parse_int(data.get("contador_total")) or parse_int(data.get("contador_pb")) or 0
            ccolor = parse_int(data.get("contador_color")) or 0
            ca3 = parse_int(data.get("contador_a3"))
            ca4 = parse_int(data.get("contador_a4"))
            ca5 = parse_int(data.get("contador_a5"))

            modelo_imp = data.get("model") or data.get("modelo") or ""

            imp, created_imp = Impressora.objects.get_or_create(
                serial_number=serial_val,
                defaults={
                    'modelo': modelo_imp,
                    'status': StatusImpressora.ESTOQUE,
                    'ultimo_contador_pb': cpb,
                    'ultimo_contador_color': ccolor,
                }
            )
            if modelo_imp and (not imp.modelo or imp.modelo != modelo_imp):
                imp.modelo = modelo_imp

            imp.ultimo_contador_pb = cpb
            imp.ultimo_contador_color = ccolor
            imp.save()

            try:
                today = timezone.localdate()
                now = timezone.now()
                hist, created_h = HistoricoContador.objects.get_or_create(
                    impressora=imp,
                    data_coleta=today,
                    origem=OrigemContador.DIARIO,
                    defaults={
                        'timestamp': now,
                        'contador_pb': cpb,
                        'contador_color': ccolor,
                        'contador_a3': ca3,
                        'contador_a4': ca4,
                        'contador_a5': ca5,
                    }
                )
                if not created_h:
                    hist.contador_pb = cpb
                    hist.contador_color = ccolor
                    if ca3 is not None: hist.contador_a3 = ca3
                    if ca4 is not None: hist.contador_a4 = ca4
                    if ca5 is not None: hist.contador_a5 = ca5
                    hist.timestamp = now
                    hist.save()
            except Exception:
                pass

        return JsonResponse({"status": "success", "message": "Métricas e histórico gravados no MySQL com sucesso"}, status=201)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def coleta_impressora_api(request):
    """
    Recebe os dados de coleta de impressora via POST e persiste no banco de dados.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método não permitido. Utilize POST."}, status=405)
        
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)
        
    # Validação simples
    ip = data.get("ip")
    status = data.get("status")
    
    if not ip or not status:
        return JsonResponse({"error": "Campos obrigatórios 'ip' e 'status' ausentes."}, status=400)
        
    serial = data.get("serial")
    
    # Validar e converter contador_geral
    contador_geral = data.get("contador_geral")
    if contador_geral is not None:
        try:
            if str(contador_geral).strip() in ("", "N/A", "---"):
                contador_geral = None
            else:
                contador_geral = int(contador_geral)
        except (ValueError, TypeError):
            contador_geral = None
            
    uptime = data.get("uptime")
    if uptime in ("N/A", "---"):
        uptime = None
        
    mensagem_painel = data.get("mensagem_painel")
    if mensagem_painel in ("N/A", "---", "Clique em Atualizar"):
        mensagem_painel = None
        
    # Validar e converter porcentagem_toner
    porcentagem_toner = data.get("porcentagem_toner")
    
    modelo = data.get("modelo")

    def parse_int_field(val):
        if val is not None:
            try:
                if str(val).strip() not in ("", "N/A", "---"):
                    return int(val)
            except (ValueError, TypeError):
                pass
        return None

    def parse_float_field(val):
        if val is not None:
            try:
                if str(val).strip() not in ("", "N/A", "---"):
                    val_str = str(val).replace("%", "").strip()
                    res = float(val_str)
                    return max(0.0, res)
            except (ValueError, TypeError):
                pass
        return None

    contador_total = parse_int_field(data.get("contador_total") or data.get("contador_geral"))
    contador_geral = parse_int_field(data.get("contador_geral") or data.get("contador_total"))
    contador_a4 = parse_int_field(data.get("contador_a4"))
    contador_a3 = parse_int_field(data.get("contador_a3"))
    contador_a5 = parse_int_field(data.get("contador_a5"))

    # Extração robusta para suprimentos aninhados ou planos
    suprimentos = data.get("suprimentos")
    if not isinstance(suprimentos, dict):
        suprimentos = {}

    tinta_preta_raw = suprimentos.get("black") if "black" in suprimentos else data.get("tinta_preta")
    tinta_ciano_raw = suprimentos.get("cyan") if "cyan" in suprimentos else data.get("tinta_ciano")
    tinta_magenta_raw = suprimentos.get("magenta") if "magenta" in suprimentos else data.get("tinta_magenta")
    tinta_amarela_raw = suprimentos.get("yellow") if "yellow" in suprimentos else data.get("tinta_amarela")
    caixa_manutencao_raw = suprimentos.get("caixa_manutencao") if "caixa_manutencao" in suprimentos else data.get("caixa_manutencao")

    tinta_preta = parse_float_field(tinta_preta_raw)
    tinta_ciano = parse_float_field(tinta_ciano_raw)
    tinta_magenta = parse_float_field(tinta_magenta_raw)
    tinta_amarela = parse_float_field(tinta_amarela_raw)
    caixa_manutencao = parse_float_field(caixa_manutencao_raw)

    if porcentagem_toner is not None:
        porcentagem_toner = parse_float_field(porcentagem_toner)
    else:
        porcentagem_toner = tinta_preta

    serial_val = serial if serial and serial not in ("N/A", "---") else f"NO-SERIAL-{ip}"

    try:
        coleta, created = ColetaImpressora.objects.update_or_create(
            serial=serial_val,
            defaults={
                "ip": ip,
                "contador_geral": contador_geral,
                "uptime": uptime,
                "mensagem_painel": mensagem_painel,
                "porcentagem_toner": porcentagem_toner,
                "status": status,
                "modelo": modelo,
                "contador_total": contador_total,
                "contador_a4": contador_a4,
                "contador_a3": contador_a3,
                "contador_a5": contador_a5,
                "tinta_preta": tinta_preta,
                "tinta_ciano": tinta_ciano,
                "tinta_magenta": tinta_magenta,
                "tinta_amarela": tinta_amarela,
                "caixa_manutencao": caixa_manutencao
            }
        )
        
        # Atualizar também o modelo Impressora e HistoricoContador (origem='DIARIO')
        if serial_val and not serial_val.startswith("NO-SERIAL"):
            cpb = contador_a4 if contador_a4 is not None else (contador_geral or contador_total or 0)
            ccolor = contador_a3 if contador_a3 is not None else (parse_int_field(data.get("contador_color")) or 0)

            imp, created_imp = Impressora.objects.get_or_create(
                serial_number=serial_val,
                defaults={
                    'modelo': modelo or '',
                    'status': StatusImpressora.ESTOQUE,
                    'ultimo_contador_pb': cpb,
                    'ultimo_contador_color': ccolor,
                }
            )
            if modelo and (not imp.modelo or imp.modelo != modelo):
                imp.modelo = modelo

            imp.ultimo_contador_pb = cpb
            imp.ultimo_contador_color = ccolor
            imp.save()

            try:
                today = timezone.localdate()
                now = timezone.now()
                hist, created_h = HistoricoContador.objects.get_or_create(
                    impressora=imp,
                    data_coleta=today,
                    origem=OrigemContador.DIARIO,
                    defaults={
                        'timestamp': now,
                        'contador_pb': cpb,
                        'contador_color': ccolor,
                        'contador_a3': contador_a3,
                        'contador_a4': contador_a4,
                        'contador_a5': contador_a5,
                    }
                )
                if not created_h:
                    hist.contador_pb = cpb
                    hist.contador_color = ccolor
                    if contador_a3 is not None: hist.contador_a3 = contador_a3
                    if contador_a4 is not None: hist.contador_a4 = contador_a4
                    if contador_a5 is not None: hist.contador_a5 = contador_a5
                    hist.timestamp = now
                    hist.save()
            except Exception:
                pass

        return JsonResponse({
            "status": "success",
            "message": "Coleta salva com sucesso no banco de dados.",
            "id": coleta.id
        }, status=201)
    except Exception as e:
        return JsonResponse({"error": f"Erro ao salvar no banco: {str(e)}"}, status=500)


from django.db import transaction

@csrf_exempt
def api_printer_search(request):
    """
    GET ou POST /api/printer/search/
    Busca a impressora pelo número de série e retorna as configurações de OIDs correspondentes.
    Se não passar 'serial', retorna a lista de todos os perfis/marcas cadastrados no banco.
    """
    if request.method == 'GET':
        serial = request.GET.get('serial') or request.GET.get('serial_number')
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            serial = data.get('serial') or data.get('serial_number')
        except Exception:
            serial = request.POST.get('serial') or request.POST.get('serial_number')
    else:
        return JsonResponse({"error": "Método não permitido."}, status=405)

    if not serial:
        # Retorna a lista de todos os perfis/marcas cadastrados
        perfis = PrinterOID.objects.select_related('brand').all()
        perfis_list = []
        for p in perfis:
            perfis_list.append({
                "marca": p.brand.name,
                "nome_perfil": p.name,
                "is_color": p.is_color,
                "is_plotter": p.is_plotter,
                "oids": {
                    "printer_oid": p.printer_oid,
                    "contador_total": p.oid_counter_total,
                    "tempo_ligada": p.oid_tempo_ligada,
                    "N_S": p.oid_serial_number,
                    "mensagem_painel": p.oid_mensagem_painel,
                    "tinta_preta": p.oid_tinta_preta,
                    "tinta_ciano": p.oid_tinta_ciano,
                    "tinta_magenta": p.oid_tinta_magenta,
                    "tinta_amarela": p.oid_tinta_amarela,
                    "caixa_manutencao": p.oid_caixa_manutencao,
                    "toner_atual": p.oid_toner_level,
                    "toner_full": p.oid_toner_full,
                    "contador_a4": p.oid_counter_mono,
                    "contador_a3": p.oid_counter_color,
                }
            })
        return JsonResponse({"status": "success", "perfis": perfis_list}, status=200)

    try:
        printer = Impressora.objects.select_related('brand', 'oid_profile').get(serial_number=serial)
    except Impressora.DoesNotExist:
        return JsonResponse({"error": "Dispositivo não cadastrado no servidor."}, status=404)

    oids_data = {}
    oid_config = printer.oid_profile or PrinterOID.objects.filter(brand=printer.brand).first() or PrinterOID.objects.filter(brands=printer.brand).first()
    if oid_config:
        oids_data = {
            "printer_oid": oid_config.printer_oid,
            "contador_total": oid_config.oid_counter_total,
            "tempo_ligada": oid_config.oid_tempo_ligada,
            "N_S": oid_config.oid_serial_number,
            "mensagem_painel": oid_config.oid_mensagem_painel,
            "tinta_preta": oid_config.oid_tinta_preta,
            "tinta_ciano": oid_config.oid_tinta_ciano,
            "tinta_magenta": oid_config.oid_tinta_magenta,
            "tinta_amarela": oid_config.oid_tinta_amarela,
            "caixa_manutencao": oid_config.oid_caixa_manutencao,
            "toner_atual": oid_config.oid_toner_level,
            "toner_full": oid_config.oid_toner_full,
            "contador_a4": oid_config.oid_counter_mono,
            "contador_a3": oid_config.oid_counter_color,
            
            # Legacy compatibility keys for unit tests
            "printer_oid": oid_config.printer_oid,
            "oid_serial_number": oid_config.oid_serial_number,
            "oid_tempo_ligada": oid_config.oid_tempo_ligada,
            "oid_mensagem_painel": oid_config.oid_mensagem_painel,
            "oid_counter_total": oid_config.oid_counter_total,
            "oid_counter_mono": oid_config.oid_counter_mono,
            "oid_counter_color": oid_config.oid_counter_color,
            "oid_toner_level": oid_config.oid_toner_level,
            "oid_toner_full": oid_config.oid_toner_full,
            "oid_tinta_preta": oid_config.oid_tinta_preta,
            "oid_tinta_ciano": oid_config.oid_tinta_ciano,
            "oid_tinta_magenta": oid_config.oid_tinta_magenta,
            "oid_tinta_amarela": oid_config.oid_tinta_amarela,
            "oid_caixa_manutencao": oid_config.oid_caixa_manutencao,
        }

    return JsonResponse({
        "serial_number": printer.serial_number,
        "marca": printer.brand.name if printer.brand else "Genérica",
        "modelo": printer.modelo or (printer.name or printer.nome_comercial or "Genérica"),
        "nome_perfil": oid_config.name if oid_config else "Padrão",
        "is_color": oid_config.is_color if oid_config else True,
        "is_plotter": oid_config.is_plotter if oid_config else False,
        "oids": oids_data
    }, status=200)


@csrf_exempt
def search_printer(request):
    """
    GET /api/printers/search/?serial_number=<NS>
    Busca a impressora por número de série. Retorna erro 404 se não existir 
    ou se não estiver com status 'Disponível no Estoque'.
    """
    if request.method != 'GET':
        return JsonResponse({"error": "Método não permitido. Utilize GET."}, status=405)
        
    serial_number = request.GET.get('serial_number')
    if not serial_number:
        return JsonResponse({"error": "Parâmetro 'serial_number' é obrigatório."}, status=400)
        
    try:
        printer = Impressora.objects.select_related('brand', 'oid_profile').get(
            serial_number=serial_number, 
            status='ESTOQUE'
        )
    except Impressora.DoesNotExist:
        return JsonResponse({"error": "Impressora não disponível ou não encontrada."}, status=404)
        
    oid_serial_number = "1.3.6.1.2.1.43.5.1.1.17.1"
    oid_profile = printer.oid_profile or PrinterOID.objects.filter(brand=printer.brand).first() or PrinterOID.objects.filter(brands=printer.brand).first()
    if oid_profile:
        oid_serial_number = oid_profile.oid_serial_number

    return JsonResponse({
        "serial_number": printer.serial_number,
        "brand": printer.brand.name if printer.brand else "Genérica",
        "contador_inicial": printer.contador_inicial,
        "status": printer.status,
        "oid_serial_number": oid_serial_number,
    }, status=200)


@csrf_exempt
def activate_printer(request):
    """
    POST /api/printers/activate/
    Recebe serial_number, name e ip_address.
    Atualiza esses dados no banco MySQL e muda o status da impressora para 'Ativa'.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método não permitido. Utilize POST."}, status=405)
        
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
        
    serial_number = data.get('serial_number')
    name = data.get('name')
    ip_address = data.get('ip_address')
    
    if not serial_number or not name or not ip_address:
        return JsonResponse({"error": "Parâmetros 'serial_number', 'name' e 'ip_address' são obrigatórios."}, status=400)
        
    try:
        with transaction.atomic():
            printer = Impressora.objects.select_for_update().get(
                serial_number=serial_number,
                status='ESTOQUE'
            )
            printer.name = name
            # Compatibilidade com campos legados
            printer.nome_comercial = name
            printer.ip_address = ip_address
            printer.status = 'ALOCADA'
            printer.data_alocacao = timezone.now()
            printer.status_sistema = 'Ativo'
            printer.save()
    except Impressora.DoesNotExist:
        return JsonResponse({"error": "Impressora não encontrada ou já ativada."}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Erro interno ao ativar a impressora: {str(e)}"}, status=500)
        
    return JsonResponse({
        "message": "Impressora ativada com sucesso!",
        "printer": {
            "serial_number": printer.serial_number,
            "name": printer.name,
            "ip_address": printer.ip_address,
            "status": printer.status,
        }
    }, status=200)


@login_required(login_url='login')
def cadastro_impressora_estoque(request):
    """
    View para gerenciar o estoque de impressoras, cadastro de marcas e OIDs.
    Renderiza o painel completo e processa os formulários.
    """
    if request.method == 'POST':
        if 'btn_cadastrar_impressora' in request.POST:
            printer_form = PrinterStockForm(request.POST)
            if printer_form.is_valid():
                printer = printer_form.save(commit=False)
                printer.status = 'ESTOQUE'
                printer.save()
                messages.success(request, f"Impressora '{printer.serial_number}' cadastrada no estoque com sucesso!")
                return redirect('cadastro_impressora_estoque')
            else:
                messages.error(request, "Erro ao cadastrar impressora. Verifique os dados inseridos.")
                brand_form = BrandForm()
                oid_form = PrinterOIDForm()
        elif 'btn_cadastrar_marca' in request.POST:
            brand_form = BrandForm(request.POST)
            if brand_form.is_valid():
                brand = brand_form.save()
                messages.success(request, f"Marca '{brand.name}' cadastrada com sucesso!")
                return redirect('cadastro_impressora_estoque')
            else:
                messages.error(request, "Erro ao cadastrar marca.")
                printer_form = PrinterStockForm()
                oid_form = PrinterOIDForm()
        elif 'btn_cadastrar_oid' in request.POST:
            oid_form = PrinterOIDForm(request.POST)
            if oid_form.is_valid():
                oid = oid_form.save()
                messages.success(request, f"OIDs configuradas com sucesso para a marca '{oid.brand.name}'!")
                return redirect('cadastro_impressora_estoque')
            else:
                messages.error(request, "Erro ao cadastrar OIDs.")
                printer_form = PrinterStockForm()
                brand_form = BrandForm()
    else:
        printer_form = PrinterStockForm()
        brand_form = BrandForm()
        oid_form = PrinterOIDForm()

    disponiveis = Impressora.objects.filter(status='ESTOQUE').select_related('brand').order_by('serial_number')
    ativas = Impressora.objects.filter(status='ALOCADA').select_related('brand').order_by('serial_number')
    brands = Brand.objects.all().order_by('name')
    oids = PrinterOID.objects.select_related('brand').all().order_by('brand__name')

    context = {
        'printer_form': printer_form,
        'brand_form': brand_form,
        'oid_form': oid_form,
        'disponiveis': disponiveis,
        'ativas': ativas,
        'brands': brands,
        'oids': oids,
    }
    return render(request, 'core/printer_cadastro.html', context)


@login_required(login_url='login')
def inventario_dashboard(request):
    """
    Dashboard de gerenciamento de inventário e logística de impressoras.
    """
    from django.db.models import Q

    clientes = Cliente.objects.all().order_by('nome')
    impressoras = Impressora.objects.all().select_related('brand', 'cliente')
    
    # Filtros por status considerando enums e valores legados
    alocadas_q = Q(status=StatusImpressora.CLIENTE) | Q(status='ALOCADA') | Q(status='Locada')
    estoque_q = Q(status=StatusImpressora.ESTOQUE) | Q(status='Disponível')
    manutencao_q = Q(status=StatusImpressora.MANUTENCAO) | Q(status='MANUTENÇÃO') | Q(status='Assistência')

    alocadas = impressoras.filter(alocadas_q).order_by('serial_number')
    estoque = impressoras.filter(estoque_q).order_by('serial_number')
    manutencao = impressoras.filter(manutencao_q).order_by('serial_number')

    total_impressoras = impressoras.count()
    total_alocadas = alocadas.count()
    total_estoque = estoque.count()
    total_manutencao = manutencao.count()
    
    context = {
        'clientes': clientes,
        'impressoras': impressoras,
        'total_impressoras': total_impressoras,
        'total_alocadas': total_alocadas,
        'total_estoque': total_estoque,
        'total_manutencao': total_manutencao,
        'estoque': estoque,
        'manutencao': manutencao,
        'alocadas': alocadas,
    }
    return render(request, 'core/inventario.html', context)


@login_required(login_url='login')
@group_required('Admin', 'Técnico')
def inventario_alocar(request):
    """
    Transiciona o status de uma impressora para CLIENTE, associando-a a um Cliente e definindo seu IP.
    """
    if request.method == 'POST':
        impressora_id = request.POST.get('impressora_id')
        cliente_id = request.POST.get('cliente_id')
        ip_address = request.POST.get('ip_address') or request.POST.get('ip_ou_hostname')
        
        impressora = get_object_or_404(Impressora, pk=impressora_id)
        cliente = get_object_or_404(Cliente, pk=cliente_id)
        
        if ip_address:
            impressora.ip_address = ip_address
            impressora.ip_ou_hostname = ip_address
            impressora.save()
            
        alterar_status_impressora(
            impressora=impressora,
            novo_status=StatusImpressora.CLIENTE,
            cliente=cliente,
            observacao=f"Alocação no cliente {cliente.nome} (IP: {ip_address or 'Não informado'})"
        )
        
        messages.success(request, f"Impressora '{impressora.serial_number}' foi alocada com sucesso no cliente '{cliente.nome}'.")
    return redirect('inventario_dashboard')


@login_required(login_url='login')
@group_required('Admin', 'Técnico')
def inventario_manutencao(request):
    """
    Envia uma impressora para o status de MANUTENÇÃO.
    """
    if request.method == 'POST':
        impressora_id = request.POST.get('impressora_id')
        descricao_problema = request.POST.get('descricao_problema') or "Reparo e verificação geral"
        
        impressora = get_object_or_404(Impressora, pk=impressora_id)
        
        from core.models import HistoricoManutencao
        HistoricoManutencao.objects.create(
            impressora=impressora,
            descricao_problema=descricao_problema
        )

        alterar_status_impressora(
            impressora=impressora,
            novo_status=StatusImpressora.MANUTENCAO,
            observacao=f"Enviada para manutenção: {descricao_problema}"
        )
        
        messages.warning(request, f"Impressora '{impressora.serial_number}' foi enviada para Manutenção.")
    return redirect('inventario_dashboard')


@login_required(login_url='login')
@group_required('Admin', 'Técnico')
def inventario_liberar(request):
    """
    Retorna uma impressora para o status de ESTOQUE.
    """
    if request.method == 'POST':
        impressora_id = request.POST.get('impressora_id')
        
        impressora = get_object_or_404(Impressora, pk=impressora_id)
        
        from core.models import HistoricoManutencao
        manutencao_aberta = HistoricoManutencao.objects.filter(impressora=impressora, data_saida__isnull=True).first()
        if manutencao_aberta:
            manutencao_aberta.data_saida = timezone.now()
            manutencao_aberta.save()
            
        alterar_status_impressora(
            impressora=impressora,
            novo_status=StatusImpressora.ESTOQUE,
            cliente=None,
            observacao="Retornou ao estoque disponível."
        )

        messages.success(request, f"Impressora '{impressora.serial_number}' retornou ao estoque.")
    return redirect('inventario_dashboard')


@csrf_exempt
def api_login_view(request):
    """
    Endpoint de login via API para autenticação do agente local.
    Retorna também a lista de clientes cadastrados ativos para vinculação.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            u = data.get('username')
            p = data.get('password')
        except Exception:
            u = request.POST.get('username')
            p = request.POST.get('password')
            
        user = authenticate(username=u, password=p)
        if user is not None:
            # Lista de clientes ativos para o agente escolher
            clientes = Cliente.objects.filter(status='Ativo').order_by('nome')
            clientes_data = [{'id': c.id, 'nome': c.nome} for c in clientes]
            return JsonResponse({
                'success': True, 
                'token': 'authenticated_session_token',
                'clientes': clientes_data
            })
        else:
            return JsonResponse({'success': False, 'error': 'Usuário ou senha incorretos.'}, status=401)
    return JsonResponse({'error': 'Método não permitido.'}, status=405)


@csrf_exempt
def api_v1_vincular_agente(request):
    """
    Endpoint POST para vincular/associar um agent_id a um Cliente específico.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido.'}, status=405)
        
    try:
        data = json.loads(request.body)
        agent_id = data.get('agent_id')
        cliente_id = data.get('cliente_id')
    except Exception:
        agent_id = request.POST.get('agent_id')
        cliente_id = request.POST.get('cliente_id')
        
    if not agent_id or not cliente_id:
        return JsonResponse({'error': "Campos 'agent_id' e 'cliente_id' são obrigatórios."}, status=400)
        
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'error': "Cliente não encontrado."}, status=404)
        
    agente, created = ComputadorAgente.objects.update_or_create(
        identificador_unico=agent_id,
        defaults={'cliente': cliente}
    )
    
    return JsonResponse({
        'success': True,
        'agent_id': agente.identificador_unico,
        'cliente_id': cliente.id,
        'cliente_nome': cliente.nome
    }, status=200)


@csrf_exempt
def api_v1_tarefas(request):
    """
    Endpoint GET para orquestração centralizada de agentes.
    Retorna apenas as impressoras alocadas vinculadas ao cliente daquele agente.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método não permitido.'}, status=405)
        
    agent_id = request.GET.get('agent_id')
    if not agent_id:
        agent_id = request.headers.get('X-Agent-ID') or request.META.get('HTTP_X_AGENT_ID')
        
    if not agent_id:
        return JsonResponse({'error': "Parâmetro 'agent_id' é obrigatório na URL ou nos headers."}, status=400)
        
    try:
        agente = ComputadorAgente.objects.select_related('cliente').get(identificador_unico=agent_id)
    except ComputadorAgente.DoesNotExist:
        return JsonResponse({'error': f"Agente com identificador '{agent_id}' não encontrado."}, status=404)
        
    if not agente.cliente:
        return JsonResponse([], safe=False, status=200)
        
    # Coleta os dados somente das impressoras que estão alocadas ("locadas") naquele cliente
    impressoras = Impressora.objects.filter(cliente=agente.cliente).filter(
        Q(status=StatusImpressora.CLIENTE) | Q(status='ALOCADA') | Q(status='CLIENTE')
    )
    
    lista_tarefas = []
    for imp in impressoras:
        ip = imp.ip_address or imp.ip_ou_hostname
        lista_tarefas.append({
            'id': imp.serial_number,
            'nome': imp.name or f"Impressora {imp.serial_number}",
            'ip': ip or "",
            'ip_ou_hostname': ip or "",
            'modelo': imp.modelo or imp.name or "Genérica",
            'marca': imp.brand.name if imp.brand else "Genérica",
            'serial_number': imp.serial_number,
            'serial_inicial': imp.serial_number,
            'perfil_oid': imp.oid_profile.name if imp.oid_profile else ""
        })
            
    return JsonResponse(lista_tarefas, safe=False, status=200)


@csrf_exempt
def api_v1_clientes_list(request):
    """
    Endpoint GET para listar todos os clientes ativos.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método não permitido.'}, status=405)
        
    clientes = Cliente.objects.filter(status='Ativo').order_by('nome')
    data = [{'id': c.id, 'nome': c.nome} for c in clientes]
    return JsonResponse(data, safe=False, status=200)


# =====================================================
# API DE COLETA E GERENCIAMENTO DE IMPRESSORAS (DRF)
# =====================================================

@csrf_exempt
def coleta_agente_api(request):
    """
    Endpoint POST para os Agentes Python enviarem dados de coleta/ping:
    - numero_serie, modelo, contador_pb, contador_color
    - Atualiza o estado atual na Impressora.
    - Executa get_or_create no HistoricoContador para o dia de hoje (origem='DIARIO').
    - Se já existir registro no dia com origem='DIARIO', apenas atualiza os contadores.
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método não permitido. Utilize POST."}, status=405)

    try:
        raw_data = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido no corpo da requisição."}, status=400)

    try:
        data = validar_coleta_agente_payload(raw_data)
    except ValueError as e:
        return JsonResponse({"error": str(e)}, status=400)

    numero_serie = data['numero_serie']
    modelo = data.get('modelo')
    contador_pb = data['contador_pb']
    contador_color = data['contador_color']

    # 1. Cadastrar/Obter a Impressora e atualizar seus valores mais recentes
    impressora, created_printer = Impressora.objects.get_or_create(
        serial_number=numero_serie,
        defaults={
            'modelo': modelo or '',
            'status': StatusImpressora.ESTOQUE,
            'ultimo_contador_pb': contador_pb,
            'ultimo_contador_color': contador_color,
        }
    )

    if modelo and (not impressora.modelo or impressora.modelo != modelo):
        impressora.modelo = modelo

    impressora.ultimo_contador_pb = contador_pb
    impressora.ultimo_contador_color = contador_color
    impressora.save()

    today = timezone.localdate()
    now = timezone.now()
    created_hist = False
    contador_mudou = created_printer

    try:
        ultimo_historico = (
            HistoricoContador.objects
            .only('id', 'impressora', 'data_coleta', 'timestamp', 'contador_pb', 'contador_color', 'origem')
            .filter(impressora=impressora)
            .order_by('-timestamp', '-id')
            .first()
        )

        contador_mudou = (
            created_printer or
            ultimo_historico is None or
            ultimo_historico.contador_pb != contador_pb or
            ultimo_historico.contador_color != contador_color
        )

        historico, created_hist = HistoricoContador.objects.get_or_create(
            impressora=impressora,
            data_coleta=today,
            origem=OrigemContador.DIARIO,
            defaults={
                'timestamp': now,
                'contador_pb': contador_pb,
                'contador_color': contador_color,
            }
        )

        if not created_hist:
            historico.contador_pb = contador_pb
            historico.contador_color = contador_color
            historico.timestamp = now
            historico.save()
    except Exception:
        pass

    return JsonResponse({
        "status": "sucesso",
        "mensagem": "Coleta processada com sucesso.",
        "impressora": serialize_impressora(impressora),
        "registro_historico_criado": created_hist,
        "contador_mudou": contador_mudou
    }, status=201 if created_printer else 200)


@csrf_exempt
def alterar_status_impressora_api(request, serial_number):
    """
    Endpoint REST para transição de status de uma impressora (ESTOQUE, CLIENTE, MANUTENCAO).
    Gera automaticamente HistoricoMovimentacao e HistoricoContador (origem='MOVIMENTACAO').
    """
    if request.method != 'POST':
        return JsonResponse({"error": "Método não permitido. Utilize POST."}, status=405)

    impressora = get_object_or_404(Impressora, serial_number=serial_number)

    try:
        payload = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido no corpo da requisição."}, status=400)

    novo_status = payload.get('status')
    cliente_id = payload.get('cliente_id')
    observacao = payload.get('observacao', '')

    if not novo_status or novo_status not in StatusImpressora.values:
        return JsonResponse({
            "error": f"Status inválido. Opções permitidas: {list(StatusImpressora.values)}"
        }, status=400)

    cliente = None
    if cliente_id:
        cliente = get_object_or_404(Cliente, pk=cliente_id)

    try:
        impressora_atualizada = alterar_status_impressora(
            impressora=impressora,
            novo_status=novo_status,
            cliente=cliente,
            observacao=observacao
        )
        return JsonResponse({
            "status": "sucesso",
            "mensagem": f"Status da impressora alterado para {novo_status}.",
            "impressora": serialize_impressora(impressora_atualizada)
        }, status=200)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


